import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# Excel template structure
TEMPLATE_COLUMNS = [
    'Plan Name',
    'Max Credits Per Semester',
    'Semester',
    'Course Name (Arabic)',
    'Course Name (English)',
    'Course Code',
    'Credit Hours',
    'Course Type',
    'Study Mode',
    'Lecture Hours',
    'Lab Hours',
    'Training Hours',
    'Department',
    'Prerequisite Codes'
]

def create_excel_template():
    """Create an Excel template for plan creation"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan Template"
        
        # Add header row with styling
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, column_title in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = [20, 25, 12, 25, 25, 15, 15, 15, 15, 15, 15, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add sample data row
        sample_data = [
            "Sample Plan", "18", "1", "مقرر عينة", "Sample Course",
            "CS101", "3", "Required", "In-Person", "3", "0", "0", "Computer Science", ""
        ]
        
        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        
        # Convert to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        print(f"Error in create_excel_template: {e}")
        raise e

def validate_excel_file(file_content):
    """Validate Excel file format and structure"""
    try:
        df = pd.read_excel(file_content)
        missing_columns = [col for col in TEMPLATE_COLUMNS if col not in df.columns]
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"
        if df.empty:
            return False, "Excel file is empty."
        return True, df
    except Exception as e:
        return False, f"Failed to read Excel file: {str(e)}"

def validate_plan_data(df):
    """Validate plan data from Excel"""
    errors = []
    plan_names = df['Plan Name'].dropna().unique()
    if len(plan_names) == 0:
        errors.append("Plan Name is required")
    elif len(plan_names) > 1:
        errors.append("All rows must have the same Plan Name")
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        if pd.isna(row['Semester']):
            errors.append(f"Row {row_num}: Semester is required")
        else:
            try:
                sem_val = int(row['Semester'])
                if not (1 <= sem_val <= 12):
                    errors.append(f"Row {row_num}: Semester must be between 1 and 12")
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Semester must be a valid number")
        if pd.isna(row['Course Name (Arabic)']): errors.append(f"Row {row_num}: Course Name (Arabic) is required")
        if pd.isna(row['Course Name (English)']): errors.append(f"Row {row_num}: Course Name (English) is required")
        if pd.isna(row['Course Code']): errors.append(f"Row {row_num}: Course Code is required")
        if pd.isna(row['Credit Hours']):
            errors.append(f"Row {row_num}: Credit Hours is required")
        else:
            try:
                cr_val = int(row['Credit Hours'])
                if cr_val <= 0:
                    errors.append(f"Row {row_num}: Credit Hours must be a positive number")
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Credit Hours must be a valid number")
    
    return errors

def export_plan_to_excel(plan_data, courses_data):
    """Export a plan to Excel format"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan Data"
        
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, column_title in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row_num = 2
        for course in courses_data:
            ws.cell(row=row_num, column=1).value = str(plan_data['name'])
            ws.cell(row=row_num, column=2).value = plan_data['max_credits_per_semester']
            ws.cell(row=row_num, column=3).value = course['semester']
            ws.cell(row=row_num, column=4).value = str(course['name_ar'])
            ws.cell(row=row_num, column=5).value = str(course['name_en'])
            ws.cell(row=row_num, column=6).value = str(course['code'])
            ws.cell(row=row_num, column=7).value = course['credits']
            ws.cell(row=row_num, column=8).value = str(course['type'])
            ws.cell(row=row_num, column=9).value = str(course['mode'])
            ws.cell(row=row_num, column=10).value = course.get('lecture_hours', 0)
            ws.cell(row=row_num, column=11).value = course.get('lab_hours', 0)
            ws.cell(row=row_num, column=12).value = course.get('training_hours', 0)
            ws.cell(row=row_num, column=13).value = str(course.get('department', ''))
            prereqs = course.get('prerequisite_codes', [])
            ws.cell(row=row_num, column=14).value = ', '.join(prereqs) if prereqs else ''
            
            for col in range(1, len(TEMPLATE_COLUMNS) + 1):
                ws.cell(row=row_num, column=col).border = border
            row_num += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        print(f"Error in export_plan_to_excel: {e}")
        raise e


# ── NEW: Alignment Excel Export ────────────────────────────────────────────────

def export_alignment_to_excel(plan_a_data, plan_b_data, alignments):
    """
    Export the alignment between two plans to a formatted Excel file (RTL).
    plan_a_data / plan_b_data: {'name': str, 'total_credits': int, 'courses_by_sem': {sem: [course, ...]}}
    alignments: list of {'course_a': course_dict, 'course_b': course_dict, 'relation_type': str}
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles import numbers as xl_numbers

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "ملخص المواءمة"
    ws_sum.sheet_view.rightToLeft = True

    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subhead_fill  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    green_fill    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    white_font    = Font(bold=True, color="FFFFFF", size=12)
    bold_font     = Font(bold=True, size=11)
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
    right_align   = Alignment(horizontal='right',  vertical='center', wrap_text=True, readingOrder=2)

    def hdr(ws, row, col, value, fill=header_fill):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = fill; c.font = white_font; c.border = thin_border; c.alignment = center_align

    def cell(ws, row, col, value, fill=None, bold=False, align=right_align):
        c = ws.cell(row=row, column=col, value=value)
        if fill: c.fill = fill
        c.font = Font(bold=bold, size=11)
        c.border = thin_border; c.alignment = align

    # Plan headers
    hdr(ws_sum, 1, 1, "البيان")
    hdr(ws_sum, 1, 2, plan_a_data['name'], subhead_fill)
    hdr(ws_sum, 1, 3, plan_b_data['name'], subhead_fill)

    rows = [
        ("إجمالي الساعات المعتمدة", plan_a_data['total_credits'], plan_b_data['total_credits']),
        ("عدد المقررات",            plan_a_data['course_count'],   plan_b_data['course_count']),
        ("عدد الأترام الدراسية",    plan_a_data['sem_count'],      plan_b_data['sem_count']),
    ]
    for r_idx, (label, val_a, val_b) in enumerate(rows, start=2):
        cell(ws_sum, r_idx, 1, label, bold=True)
        cell(ws_sum, r_idx, 2, val_a, align=center_align)
        cell(ws_sum, r_idx, 3, val_b, align=center_align)

    # Smart cards using Excel formulas — alignment stats
    hdr(ws_sum, 6, 1, "إحصائيات المواءمة", header_fill)
    ws_sum.merge_cells('B6:C6')
    ws_sum.cell(6, 2).fill = header_fill; ws_sum.cell(6, 2).font = white_font

    total_aligned = len(alignments)
    total_a       = plan_a_data['course_count']
    pct           = round(total_aligned / total_a * 100, 1) if total_a else 0

    card_rows = [
        ("عدد الأزواج المترابطة",       total_aligned),
        ("نسبة المواءمة من خطة (أ)",    f"{pct}%"),
        ("المواد غير المواءمة في (أ)",  total_a - total_aligned),
        ("المواد غير المواءمة في (ب)",  plan_b_data['course_count'] - total_aligned),
    ]
    for r_idx, (label, val) in enumerate(card_rows, start=7):
        cell(ws_sum, r_idx, 1, label, bold=True)
        ws_sum.merge_cells(f'B{r_idx}:C{r_idx}')
        cell(ws_sum, r_idx, 2, val, align=center_align)

    ws_sum.column_dimensions['A'].width = 32
    ws_sum.column_dimensions['B'].width = 26
    ws_sum.column_dimensions['C'].width = 26
    ws_sum.row_dimensions[1].height = 30

    # ── Sheet 2: Mapping Details ──────────────────────────────────────────────
    ws_map = wb.create_sheet("تفاصيل المواءمة")
    ws_map.sheet_view.rightToLeft = True

    map_headers = [
        "رمز المقرر (أ)", "اسم المقرر (أ)", "ساعات (أ)", "الترم (أ)",
        "نوع العلاقة",
        "رمز المقرر (ب)", "اسم المقرر (ب)", "ساعات (ب)", "الترم (ب)",
        "الحالة"
    ]
    for col, h in enumerate(map_headers, 1):
        hdr(ws_map, 1, col, h)

    accepted_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    review_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    rel_labels = {'equivalent': 'مكافئ', 'similar': 'مشابه', 'replaced_by': 'يحل محله'}

    for r_idx, aln in enumerate(alignments, start=2):
        ca = aln['course_a']
        cb = aln['course_b']
        rel = aln['relation_type']
        status = "معتمد تلقائياً" if rel == 'equivalent' else "يتطلب مراجعة"
        row_fill = accepted_fill if rel == 'equivalent' else review_fill

        vals = [
            ca['code'], ca['name_ar'], ca['credits'], ca['semester'],
            rel_labels.get(rel, rel),
            cb['code'], cb['name_ar'], cb['credits'], cb['semester'],
            status
        ]
        for col, v in enumerate(vals, 1):
            c = ws_map.cell(row=r_idx, column=col, value=v)
            c.fill = row_fill; c.border = thin_border; c.alignment = right_align

    col_widths = [16, 28, 10, 10, 14, 16, 28, 10, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws_map.column_dimensions[get_column_letter(i)].width = w
    ws_map.row_dimensions[1].height = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
